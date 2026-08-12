"""Views de estadísticas con exportación a PDF, CSV y Excel.

Acceso:
  - ADMINISTRADOR, ENCARGADO, AUDITOR → todas las vistas estadísticas y exportación
"""
import csv
import io
import os
from datetime import date, timedelta
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import BasePermission
from django.db import connection
from django.http import HttpResponse
from django.conf import settings
from api.permissions import get_user_role
from api.utils.auditoria import log_descarga


class IsStatsViewer(BasePermission):
    """ADMINISTRADOR, ENCARGADO, DIRECTOR_SERVICIO_MEDICO y AUDITOR pueden ver y exportar estadísticas."""
    def has_permission(self, request, view):
        return get_user_role(request.user) in ('ADMINISTRADOR', 'ENCARGADO', 'AUDITOR', 'DIRECTOR_SERVICIO_MEDICO')


def dictfetchall(cursor):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _get_schemas(request):
    dep = request.query_params.get('departamento', 'ambos').lower()
    if dep == 'farmacia': return ['farmacia']
    elif dep in ('proveeduria', 'proveeduria'): return ['proveeduria']
    return ['farmacia', 'proveeduria']

def _get_valid_schemas(schemas, table_alias):
    """Devuelve solo los esquemas que contienen la tabla o vista requerida."""
    valid = []
    for sch in schemas:
        if sch == 'farmacia':
            valid.append(sch)
        elif sch == 'proveeduria':
            # Proveeduría aún no implementa despachos_actas ni inventario_detalle de forma homóloga
            if table_alias in ('lotes',):
                valid.append(sch)
    return valid

def _get_union_tables(schemas, table_alias):
    """Construye un UNION ALL seguro combinando y adaptando esquemas si es necesario."""
    valid = _get_valid_schemas(schemas, table_alias)
    if not valid:
        return f"SELECT 'farmacia' as schema_name, * FROM farmacia.{table_alias} WHERE 1=0"
    return " UNION ALL ".join([f"SELECT '{sch}' as schema_name, * FROM {sch}.{table_alias}" for sch in valid])

def _get_despachos_tables(schemas):
    queries = []
    for sch in schemas:
        if sch == 'farmacia':
            queries.append("SELECT 'farmacia' as schema_name, da.orden_id::text AS id_acta, da.folio_grupo::text AS folio_grupo, da.fecha_hora, da.cantidad_despachada, l.id_med_base FROM farmacia.despachos_actas da JOIN farmacia.lotes l ON da.id_lote = l.id_lote")
        elif sch == 'proveeduria':
            queries.append("SELECT 'proveeduria' as schema_name, sd.id_detalle::text AS id_acta, s.folio_solicitud::text AS folio_grupo, s.fecha_entrega AS fecha_hora, sd.cantidad_entregada AS cantidad_despachada, sd.id_med_base FROM proveeduria.solicitudes_detalle sd JOIN proveeduria.solicitudes s ON sd.id_solicitud = s.id_solicitud WHERE s.estado = 'ENTREGADA'")
    if not queries: return "SELECT 'farmacia' as schema_name, '1' as id_acta, '' as folio_grupo, CURRENT_TIMESTAMP as fecha_hora, 0 as cantidad_despachada, 1 as id_med_base WHERE 1=0"
    return " UNION ALL ".join(queries)

def _get_inventario_tables(schemas):
    queries = []
    for sch in schemas:
        if sch == 'farmacia':
            queries.append("SELECT 'farmacia' as schema_name, v.estado_logico, v.cantidad_actual, v.activo, v.id_lote FROM farmacia.inventario_detalle v")
        elif sch == 'proveeduria':
            queries.append("SELECT 'proveeduria' as schema_name, CASE WHEN l.fecha_vencimiento < CURRENT_DATE THEN 'VENCIDO' WHEN l.cantidad_actual <= 0 THEN 'AGOTADO' WHEN l.fecha_vencimiento <= (CURRENT_DATE + INTERVAL '90 days') THEN 'PRÓXIMO A VENCER' ELSE 'ÓPTIMO' END as estado_logico, l.cantidad_actual, l.activo, l.id_lote FROM proveeduria.lotes l")
    if not queries: return "SELECT 'farmacia' as schema_name, 'ÓPTIMO' as estado_logico, 0 as cantidad_actual, FALSE as activo, 1 as id_lote WHERE 1=0"
    return " UNION ALL ".join(queries)

def get_date_range(request):
    """Extrae y valida el rango de fechas del request."""
    hoy = date.today()
    fecha_desde = request.query_params.get('desde', str(hoy - timedelta(days=30)))
    fecha_hasta = request.query_params.get('hasta', str(hoy))
    return fecha_desde, fecha_hasta


class EstadisticasResumenView(APIView):
    """Resumen de dotaciones vs despachos en el período."""
    permission_classes = [IsStatsViewer]

    def get(self, request):
        fecha_desde, fecha_hasta = get_date_range(request)
        modo = request.query_params.get('modo', 'medicamentos').lower()
        
        schemas = _get_schemas(request)
        lotes_tables = _get_union_tables(schemas, 'lotes')
        despachos_tables = _get_despachos_tables(schemas)
        
        with connection.cursor() as cursor:
            # 1. DOTACIONES (Ingresos en el período)
            if modo == 'medicamentos':
                sql_dot = f"""
                    SELECT COUNT(DISTINCT CONCAT(schema_name, id_med_base)) AS total, COALESCE(SUM(cantidad_inicial), 0) AS unidades
                    FROM ({lotes_tables}) lotes
                    WHERE DATE(fecha_ingreso) BETWEEN %s AND %s
                """
            else:
                sql_dot = f"""
                    SELECT COUNT(*) AS total, COALESCE(SUM(cantidad_inicial), 0) AS unidades
                    FROM ({lotes_tables}) lotes
                    WHERE DATE(fecha_ingreso) BETWEEN %s AND %s
                """
            cursor.execute(sql_dot, [fecha_desde, fecha_hasta])
            dot = cursor.fetchone()

            # 2. DESPACHOS (Salidas en el período)
            # Siempre traemos ambos: variedad de medicamentos y total de actas (transacciones)
            cursor.execute(f"""
                SELECT 
                    COUNT(DISTINCT CONCAT(da.schema_name, da.id_med_base)) AS variedad, 
                    COUNT(DISTINCT da.folio_grupo) AS transacciones,
                    COALESCE(SUM(da.cantidad_despachada), 0) AS unidades
                FROM ({despachos_tables}) da
                
                WHERE DATE(da.fecha_hora) BETWEEN %s AND %s
            """, [fecha_desde, fecha_hasta])
            des = cursor.fetchone()

            # 3. INVENTARIO ACTUAL (Estado actual del sistema)
            if modo == 'medicamentos':
                sql_inv = f"""
                    SELECT COUNT(DISTINCT CONCAT(schema_name, id_med_base)) AS total, COALESCE(SUM(cantidad_actual), 0) AS unidades
                    FROM ({lotes_tables}) lotes
                    WHERE activo = TRUE AND cantidad_actual > 0
                """
            else:
                sql_inv = f"""
                    SELECT COUNT(*) AS total, COALESCE(SUM(cantidad_actual), 0) AS unidades
                    FROM ({lotes_tables}) lotes
                    WHERE activo = TRUE AND cantidad_actual > 0
                """
            cursor.execute(sql_inv)
            inv = cursor.fetchone()

        return Response({
            'periodo': f'{fecha_desde} al {fecha_hasta}',
            'modo': modo,
            'dotaciones_count': dot[0],
            'dotaciones_unidades': int(dot[1]),
            'despachos_count': des[0], # Variedad
            'despachos_transacciones': des[1], # Total Actas
            'despachos_unidades': int(des[2]),
            'inventario_count': inv[0],
            'inventario_unidades': int(inv[1])
        })


class EstadoInventarioChartView(APIView):
    """Distribución actual del inventario por estado para gráfica circular."""
    permission_classes = [IsStatsViewer]

    def get(self, request):
        modo = request.query_params.get('modo', 'medicamentos').lower()
        fecha_desde, fecha_hasta = get_date_range(request)
        
        # Definir los 4 estados base con sus colores para asegurar que siempre aparezcan
        stats_map = {
            'ÓPTIMO': {'color': '#25b003', 'cantidad': 0},
            'PRÓXIMO A VENCER': {'color': '#e5780b', 'cantidad': 0},
            'VENCIDO': {'color': '#e51a1a', 'cantidad': 0},
            'AGOTADO': {'color': '#6c757d', 'cantidad': 0},
        }

        schemas = _get_schemas(request)
        inventario_tables = _get_inventario_tables(schemas)
        lotes_tables = _get_union_tables(schemas, 'lotes')

        with connection.cursor() as cursor:
            # Para el "Balance de Inventario", lo correcto es ver el estado acumulado
            # hasta la fecha final seleccionada (snapshot histórico).
            if modo == 'medicamentos':
                # Cantidad Total de unidades físicas acumuladas hasta fecha_hasta
                cursor.execute(f"""
                    SELECT 
                        v.estado_logico, 
                        SUM(v.cantidad_actual) AS cantidad
                    FROM ({inventario_tables}) v
                    JOIN ({lotes_tables}) l ON v.id_lote = l.id_lote
                    WHERE v.activo = TRUE AND DATE(l.fecha_ingreso) <= %s
                    GROUP BY v.estado_logico
                """, [fecha_hasta])
            else:
                # Conteo de lotes acumulados hasta fecha_hasta
                cursor.execute(f"""
                    SELECT 
                        v.estado_logico, 
                        COUNT(*) AS cantidad
                    FROM ({inventario_tables}) v
                    JOIN ({lotes_tables}) l ON v.id_lote = l.id_lote
                    WHERE v.activo = TRUE AND DATE(l.fecha_ingreso) <= %s
                    GROUP BY v.estado_logico
                """, [fecha_hasta])
            
            rows = cursor.fetchall()
            
            # Llenar el mapa con los resultados reales
            for r in rows:
                estado = r[0]
                if estado in stats_map:
                    stats_map[estado]['cantidad'] = float(r[1])

        # Construir respuesta en orden fijo para el frontend
        orden = ['ÓPTIMO', 'PRÓXIMO A VENCER', 'VENCIDO', 'AGOTADO']
        data = []
        total = sum(item['cantidad'] for item in stats_map.values())
        
        for estado in orden:
            item = stats_map[estado]
            data.append({
                'estado': estado,
                'color_clase': item['color'],
                'cantidad': item['cantidad'],
                'porcentaje': round(item['cantidad'] / total * 100, 1) if total > 0 else 0
            })

        return Response(data)


class InventarioPorCategoriaView(APIView):
    """Distribución actual del inventario agrupado por categoría."""
    permission_classes = [IsStatsViewer]

    def get(self, request):
        fecha_desde, fecha_hasta = get_date_range(request)
        
        schemas = _get_schemas(request)
        lotes_tables = _get_union_tables(schemas, 'lotes')

        with connection.cursor() as cursor:
            cursor.execute(f"""
                SELECT 
                    COALESCE(c.nombre_categoria, 'Sin Categoría') AS categoria,
                    COUNT(DISTINCT CONCAT(l.schema_name, m.id_med_base)) AS variedad,
                    SUM(l.cantidad_inicial) AS total_unidades
                FROM ({lotes_tables}) l
                JOIN farmacia.medicamentos_base m ON l.id_med_base = m.id_med_base
                LEFT JOIN farmacia.categorias_medicamento c ON m.id_categoria = c.id_categoria
                WHERE DATE(l.fecha_ingreso) BETWEEN %s AND %s
                GROUP BY categoria
                ORDER BY total_unidades DESC
            """, [fecha_desde, fecha_hasta])
            rows = cursor.fetchall()
            
            # Si no hay movimientos en el periodo, podemos mostrar el existencia actual para no dejarlo vacío
            if not rows:
                cursor.execute(f"""
                    SELECT 
                        COALESCE(c.nombre_categoria, 'Sin Categoría') AS categoria,
                        COUNT(DISTINCT CONCAT(l.schema_name, m.id_med_base)) AS variedad,
                        SUM(l.cantidad_actual) AS total_unidades
                    FROM ({lotes_tables}) l
                    JOIN farmacia.medicamentos_base m ON l.id_med_base = m.id_med_base
                    LEFT JOIN farmacia.categorias_medicamento c ON m.id_categoria = c.id_categoria
                    WHERE l.activo = TRUE AND l.cantidad_actual > 0
                    GROUP BY categoria
                    ORDER BY total_unidades DESC
                """)
                rows = cursor.fetchall()
        
        return Response([
            {'categoria': r[0], 'variedad': r[1], 'unidades': int(r[2])}
            for r in rows
        ])


class DespachosPorMedicamentoView(APIView):
    """Top medicamentos más despachados en el período."""
    permission_classes = [IsStatsViewer]

    def get(self, request):
        fecha_desde, fecha_hasta = get_date_range(request)
        limit = int(request.query_params.get('limit', 10))

        schemas = _get_schemas(request)
        lotes_tables = _get_union_tables(schemas, 'lotes')
        despachos_tables = _get_despachos_tables(schemas)

        with connection.cursor() as cursor:
            cursor.execute(f"""
                SELECT mb.nombre_generico, SUM(da.cantidad_despachada) AS total_despachado
                FROM ({despachos_tables}) da
                JOIN farmacia.medicamentos_base mb ON da.id_med_base = mb.id_med_base
                WHERE DATE(da.fecha_hora) BETWEEN %s AND %s
                GROUP BY mb.nombre_generico
                ORDER BY total_despachado DESC
                LIMIT %s
            """, [fecha_desde, fecha_hasta, limit])
            rows = cursor.fetchall()

        total = sum(r[1] for r in rows)
        data = [
            {
                'nombre_generico': r[0],
                'total_despachado': int(r[1]),
                'porcentaje': round(r[1] / total * 100, 1) if total > 0 else 0
            }
            for r in rows
        ]
        return Response(data)


class EvolucionTemporalView(APIView):
    """Evolución de dotaciones y despachos por día en el período."""
    permission_classes = [IsStatsViewer]

    def get(self, request):
        fecha_desde, fecha_hasta = get_date_range(request)
        modo = request.query_params.get('modo', 'medicamentos').lower()

        schemas = _get_schemas(request)
        lotes_tables = _get_union_tables(schemas, 'lotes')
        despachos_tables = _get_despachos_tables(schemas)

        with connection.cursor() as cursor:
            # Dotaciones
            if modo == 'medicamentos':
                cursor.execute(f"""
                    SELECT DATE(fecha_ingreso) AS fecha, COUNT(DISTINCT CONCAT(schema_name, id_med_base)) AS total, COALESCE(SUM(cantidad_inicial),0) AS unidades
                    FROM ({lotes_tables}) lotes
                    WHERE DATE(fecha_ingreso) BETWEEN %s AND %s
                    GROUP BY DATE(fecha_ingreso) ORDER BY fecha
                """, [fecha_desde, fecha_hasta])
            else:
                cursor.execute(f"""
                    SELECT DATE(fecha_ingreso) AS fecha, COUNT(*) AS total, COALESCE(SUM(cantidad_inicial),0) AS unidades
                    FROM ({lotes_tables}) lotes
                    WHERE DATE(fecha_ingreso) BETWEEN %s AND %s
                    GROUP BY DATE(fecha_ingreso) ORDER BY fecha
                """, [fecha_desde, fecha_hasta])
            dotaciones = {str(r[0]): {'total': r[1], 'unidades': int(r[2])} for r in cursor.fetchall()}

            # Despachos
            if modo == 'medicamentos':
                cursor.execute(f"""
                    SELECT DATE(da.fecha_hora) AS fecha, COUNT(DISTINCT CONCAT(da.schema_name, da.id_med_base)) AS total, COALESCE(SUM(da.cantidad_despachada),0) AS unidades
                    FROM ({despachos_tables}) da
                    
                    WHERE DATE(da.fecha_hora) BETWEEN %s AND %s
                    GROUP BY DATE(da.fecha_hora) ORDER BY fecha
                """, [fecha_desde, fecha_hasta])
            else:
                cursor.execute(f"""
                    SELECT DATE(fecha_hora) AS fecha, COUNT(*) AS total, COALESCE(SUM(cantidad_despachada),0) AS unidades
                    FROM ({despachos_tables}) despachos_actas
                    WHERE DATE(fecha_hora) BETWEEN %s AND %s
                    GROUP BY DATE(fecha_hora) ORDER BY fecha
                """, [fecha_desde, fecha_hasta])
            despachos = {str(r[0]): {'total': r[1], 'unidades': int(r[2])} for r in cursor.fetchall()}

        # Combinar por fechas
        all_dates = sorted(set(list(dotaciones.keys()) + list(despachos.keys())))
        data = []
        for d in all_dates:
            dot = dotaciones.get(d, {'total': 0, 'unidades': 0})
            des = despachos.get(d, {'total': 0, 'unidades': 0})
            data.append({
                'fecha': d,
                'dotaciones_cantidad': dot['total'],
                'dotaciones_unidades': dot['unidades'],
                'despachos_cantidad': des['total'],
                'despachos_unidades': des['unidades'],
            })
        return Response(data)


class ExportarEstadisticasView(APIView):
    """Exporta estadísticas en CSV, Excel o PDF."""
    permission_classes = [IsStatsViewer]

    def get(self, request):
        formato = request.query_params.get('formato', 'csv').lower()
        tipo = request.query_params.get('tipo', 'despachos').lower()
        fecha_desde, fecha_hasta = get_date_range(request)

        fecha_desde_fmt = fecha_desde[8:10] + '-' + fecha_desde[5:7] + '-' + fecha_desde[0:4]
        fecha_hasta_fmt = fecha_hasta[8:10] + '-' + fecha_hasta[5:7] + '-' + fecha_hasta[0:4]

        schemas = _get_schemas(request)
        with connection.cursor() as cursor:
            if tipo == 'inventario':
                q_list = []
                for sch in schemas:
                    q_list.append(f"""
                        SELECT 
                            l.numero_lote, mb.nombre_generico, pm.nombre_presentacion,
                            (
                                SELECT string_agg(UPPER(pa.nombre_principio || ' ' || mc.concentracion_valor || ' ' || u.nombre_unidad), ' - ')
                                FROM farmacia.medicamento_componentes mc
                                JOIN farmacia.principios_activos pa ON mc.id_principio = pa.id_principio
                                JOIN farmacia.unidades_medida u ON mc.id_unidad = u.id_unidad
                                WHERE mc.id_med_base = mb.id_med_base
                            ) AS componentes,
                            l.cantidad_actual, TO_CHAR(l.fecha_vencimiento, 'DD-MM-YYYY'),
                            CASE 
                                WHEN l.fecha_vencimiento < CURRENT_DATE THEN 'VENCIDO'
                                WHEN l.cantidad_actual <= 0 THEN 'AGOTADO'
                                WHEN l.fecha_vencimiento <= CURRENT_DATE + INTERVAL '90 days' THEN 'PRÓXIMO A VENCER'
                                ELSE 'ÓPTIMO'
                            END AS estado,
                            CASE WHEN '{sch}' = 'proveeduria' THEN 'PROVEEDURIA' ELSE UPPER('{sch}') END AS departamento
                        FROM {sch}.lotes l
                        JOIN farmacia.medicamentos_base mb ON l.id_med_base = mb.id_med_base
                        LEFT JOIN farmacia.presentaciones_medicamento pm ON mb.id_presentacion = pm.id_presentacion
                    """)
                
                sql = " UNION ALL ".join(q_list) + " ORDER BY 6 ASC"
                cursor.execute(sql)
                rows = cursor.fetchall()
                columns = ['Lote', 'Medicamento', 'Presentación', 'Componentes', 'Cantidad Actual', 'Fecha Vencimiento', 'Estado', 'Departamento']
                titulo = "Estado Actual del Inventario"
                hoy_str = date.today().strftime('%d-%m-%Y')
                filename = f"inventario_{hoy_str}"
                
            elif tipo == 'top_medicamentos':
                limit = int(request.query_params.get('limit', 100))
                q_list = []
                for sch in schemas:
                    if sch == 'farmacia':
                        q_list.append(f"""
                            SELECT 
                                mb.id_med_base, mb.nombre_generico, COALESCE(pm.nombre_presentacion,'N/A') AS presentacion,
                                (
                                    SELECT string_agg(UPPER(pa.nombre_principio || ' ' || mc.concentracion_valor || ' ' || u.nombre_unidad), ' - ')
                                    FROM farmacia.medicamento_componentes mc
                                    JOIN farmacia.principios_activos pa ON mc.id_principio = pa.id_principio
                                    JOIN farmacia.unidades_medida u ON mc.id_unidad = u.id_unidad
                                    WHERE mc.id_med_base = mb.id_med_base
                                ) AS componentes,
                                SUM(da.cantidad_despachada) AS total_despachado,
                                'FARMACIA' AS departamento
                            FROM farmacia.despachos_actas da
                            JOIN farmacia.lotes l ON da.id_lote = l.id_lote
                            JOIN farmacia.medicamentos_base mb ON l.id_med_base = mb.id_med_base
                            LEFT JOIN farmacia.presentaciones_medicamento pm ON mb.id_presentacion = pm.id_presentacion
                            WHERE DATE(da.fecha_hora) BETWEEN %s AND %s
                            GROUP BY mb.id_med_base, mb.nombre_generico, presentacion
                        """)
                    elif sch == 'proveeduria':
                        q_list.append(f"""
                            SELECT 
                                mb.id_med_base, mb.nombre_generico, COALESCE(pm.nombre_presentacion,'N/A') AS presentacion,
                                (
                                    SELECT string_agg(UPPER(pa.nombre_principio || ' ' || mc.concentracion_valor || ' ' || u.nombre_unidad), ' - ')
                                    FROM farmacia.medicamento_componentes mc
                                    JOIN farmacia.principios_activos pa ON mc.id_principio = pa.id_principio
                                    JOIN farmacia.unidades_medida u ON mc.id_unidad = u.id_unidad
                                    WHERE mc.id_med_base = mb.id_med_base
                                ) AS componentes,
                                SUM(sd.cantidad_entregada) AS total_despachado,
                                'PROVEEDURIA' AS departamento
                            FROM proveeduria.solicitudes_detalle sd
                            JOIN proveeduria.solicitudes s ON sd.id_solicitud = s.id_solicitud
                            JOIN farmacia.medicamentos_base mb ON sd.id_med_base = mb.id_med_base
                            LEFT JOIN farmacia.presentaciones_medicamento pm ON mb.id_presentacion = pm.id_presentacion
                            WHERE s.estado = 'ENTREGADA' AND DATE(s.fecha_entrega) BETWEEN %s AND %s
                            GROUP BY mb.id_med_base, mb.nombre_generico, presentacion
                        """)
                if not q_list:
                    rows = []
                else:
                    sql = f"SELECT nombre_generico, presentacion, MAX(componentes), SUM(total_despachado) as total, departamento FROM ( {' UNION ALL '.join(q_list)} ) sub GROUP BY id_med_base, nombre_generico, presentacion, departamento ORDER BY total DESC LIMIT %s"
                    params = [fecha_desde, fecha_hasta] * len(schemas) + [limit]
                    cursor.execute(sql, params)
                    rows = cursor.fetchall()
                columns = ['Medicamento', 'Presentación', 'Componentes', 'Total Despachado', 'Departamento']
                titulo = f"Medicamentos más despachados — {fecha_desde_fmt} al {fecha_hasta_fmt}"
                filename = f"top_medicamentos_{fecha_desde_fmt}_{fecha_hasta_fmt}"

            elif tipo == 'ingresos':
                q_list = []
                for sch in schemas:
                    q_list.append(f"""
                        SELECT 
                            l.numero_lote, mb.nombre_generico, COALESCE(pm.nombre_presentacion,'N/A') AS presentacion,
                            (
                                SELECT string_agg(UPPER(pa.nombre_principio || ' ' || mc.concentracion_valor || ' ' || u.nombre_unidad), ' - ')
                                FROM farmacia.medicamento_componentes mc
                                JOIN farmacia.principios_activos pa ON mc.id_principio = pa.id_principio
                                JOIN farmacia.unidades_medida u ON mc.id_unidad = u.id_unidad
                                WHERE mc.id_med_base = mb.id_med_base
                            ) AS componentes,
                            l.cantidad_inicial, 
                            TO_CHAR(((l.fecha_ingreso AT TIME ZONE 'UTC') AT TIME ZONE 'America/Caracas'), 'DD-MM-YYYY HH24:MI') AS fecha,
                            COALESCE(au.username, 'Sistema') AS usuario,
                            CASE WHEN al.descripcion ILIKE '%%Carga masiva%%' THEN 'MASIVO' ELSE 'MANUAL' END as tipo_carga,
                            CASE WHEN '{sch}' = 'proveeduria' THEN 'PROVEEDURIA' ELSE UPPER('{sch}') END AS departamento,
                            l.fecha_ingreso
                        FROM {sch}.lotes l
                        JOIN farmacia.medicamentos_base mb ON l.id_med_base = mb.id_med_base
                        LEFT JOIN farmacia.presentaciones_medicamento pm ON mb.id_presentacion = pm.id_presentacion
                        LEFT JOIN public.auth_user au ON l.usuario_registro = au.id
                        LEFT JOIN public.auditoria_logs al ON al.accion IN ('INCLUSION', 'INVENTARIO_INCLUSION') 
                             AND al.descripcion ILIKE '%%' || l.numero_lote || '%%'
                        WHERE DATE((l.fecha_ingreso AT TIME ZONE 'UTC') AT TIME ZONE 'America/Caracas') BETWEEN %s AND %s
                    """)
                sql = f"SELECT numero_lote, nombre_generico, presentacion, componentes, cantidad_inicial, fecha, usuario, tipo_carga, departamento FROM ( {' UNION ALL '.join(q_list)} ) sub ORDER BY fecha_ingreso DESC"
                params = [fecha_desde, fecha_hasta] * len(schemas)
                cursor.execute(sql, params)
                rows = cursor.fetchall()
                columns = ['Lote', 'Medicamento', 'Presentación', 'Componentes', 'Cantidad Inicial', 'Fecha Ingreso', 'Usuario', 'Tipo Carga', 'Departamento']
                titulo = f"Auditoría de Ingresos/Dotaciones — {fecha_desde_fmt} al {fecha_hasta_fmt}"
                filename = f"ingresos_{fecha_desde_fmt}_{fecha_hasta_fmt}"
                
            else: # despachos
                q_list = []
                for sch in schemas:
                    if sch == 'farmacia':
                        q_list.append(f"""
                            SELECT
                                UPPER(LEFT(da.folio_grupo::text, 8)) as orden_id, 
                                TO_CHAR(da.fecha_hora, 'DD-MM-YYYY HH24:MI') as fecha_hora, 
                                mb.nombre_generico,
                                COALESCE(pm.nombre_presentacion,'N/A') AS presentacion,
                                (
                                    SELECT string_agg(UPPER(pa.nombre_principio || ' ' || mc.concentracion_valor || ' ' || u.nombre_unidad), ' - ')
                                    FROM farmacia.medicamento_componentes mc
                                    JOIN farmacia.principios_activos pa ON mc.id_principio = pa.id_principio
                                    JOIN farmacia.unidades_medida u ON mc.id_unidad = u.id_unidad
                                    WHERE mc.id_med_base = mb.id_med_base
                                ) AS componentes,
                                l.numero_lote, da.cantidad_despachada,
                                da.cedula_beneficiario::text AS cedula_beneficiario, da.es_carga,
                                COALESCE(
                                    t_por_contacto.nombres || ' ' || t_por_contacto.apellidos,
                                    t2.nombres || ' ' || t2.apellidos,
                                    t1.nombres || ' ' || t1.apellidos,
                                    da.nombre_beneficiario
                                ) AS titular_nombre,
                                COALESCE(
                                    t_por_contacto.cedula,
                                    t2.cedula,
                                    t1.cedula,
                                    da.cedula_beneficiario::text
                                ) AS titular_cedula,
                                COALESCE(au.first_name||' '||au.last_name,'Sistema') AS farmaceuta,
                                da.fecha_hora as sort_date
                            FROM farmacia.despachos_actas da
                            JOIN farmacia.lotes l ON da.id_lote = l.id_lote
                            JOIN farmacia.medicamentos_base mb ON l.id_med_base = mb.id_med_base
                            LEFT JOIN farmacia.presentaciones_medicamento pm ON mb.id_presentacion = pm.id_presentacion
                            LEFT JOIN public.auth_user au ON da.id_usuario_farmaceuta = au.id
                            LEFT JOIN public.administrador_titular t1 ON 
                                REGEXP_REPLACE(t1.cedula, '[^0-9]', '', 'g') = REGEXP_REPLACE(da.cedula_beneficiario::text, '[^0-9]', '', 'g')
                            LEFT JOIN public.administrador_cargafamiliar cf ON 
                                REGEXP_REPLACE(cf.cedula, '[^0-9]', '', 'g') = REGEXP_REPLACE(da.cedula_beneficiario::text, '[^0-9]', '', 'g')
                            LEFT JOIN public.administrador_titular t2 ON 
                                REGEXP_REPLACE(cf.titular_id, '[^0-9]', '', 'g') = REGEXP_REPLACE(t2.cedula, '[^0-9]', '', 'g')
                            LEFT JOIN public.administrador_titular t_por_contacto ON (
                                (da.correo_beneficiario IS NOT NULL AND da.correo_beneficiario <> '' AND t_por_contacto.correo = da.correo_beneficiario)
                                OR 
                                (da.telefono_beneficiario IS NOT NULL AND da.telefono_beneficiario <> '' AND t_por_contacto.telefono_principal = da.telefono_beneficiario)
                            ) AND da.es_carga = true
                            WHERE DATE(da.fecha_hora) BETWEEN %s AND %s
                        """)
                    elif sch == 'proveeduria':
                        q_list.append(f"""
                            SELECT
                                UPPER(LEFT(s.folio_solicitud::text, 8)) as orden_id, 
                                TO_CHAR(s.fecha_entrega, 'DD-MM-YYYY HH24:MI') as fecha_hora, 
                                mb.nombre_generico,
                                COALESCE(pm.nombre_presentacion,'N/A') AS presentacion,
                                (
                                    SELECT string_agg(UPPER(pa.nombre_principio || ' ' || mc.concentracion_valor || ' ' || u.nombre_unidad), ' - ')
                                    FROM farmacia.medicamento_componentes mc
                                    JOIN farmacia.principios_activos pa ON mc.id_principio = pa.id_principio
                                    JOIN farmacia.unidades_medida u ON mc.id_unidad = u.id_unidad
                                    WHERE mc.id_med_base = mb.id_med_base
                                ) AS componentes,
                                'N/A' AS numero_lote,
                                sd.cantidad_entregada AS cantidad_despachada,
                                'N/A' AS cedula_beneficiario, FALSE AS es_carga,
                                s.destino AS titular_nombre,
                                'N/A' AS titular_cedula,
                                COALESCE(au.first_name||' '||au.last_name,'Sistema') AS farmaceuta,
                                s.fecha_entrega as sort_date
                            FROM proveeduria.solicitudes_detalle sd
                            JOIN proveeduria.solicitudes s ON sd.id_solicitud = s.id_solicitud
                            JOIN farmacia.medicamentos_base mb ON sd.id_med_base = mb.id_med_base
                            LEFT JOIN farmacia.presentaciones_medicamento pm ON mb.id_presentacion = pm.id_presentacion
                            LEFT JOIN public.auth_user au ON s.id_usuario_procesador = au.id
                            WHERE s.estado = 'ENTREGADA' AND DATE(s.fecha_entrega) BETWEEN %s AND %s
                        """)
                if not q_list:
                    rows = []
                else:
                    sql = f"SELECT orden_id, fecha_hora, nombre_generico, presentacion, componentes, numero_lote, cantidad_despachada, COALESCE(NULLIF(cedula_beneficiario::text, ''), titular_cedula::text) as cedula_beneficiario, es_carga, titular_nombre, titular_cedula, farmaceuta FROM ( {' UNION ALL '.join(q_list)} ) sub ORDER BY sort_date DESC"
                    params = [fecha_desde, fecha_hasta] * len(schemas)
                    cursor.execute(sql, params)
                    rows = cursor.fetchall()
                columns = ['Nº Acta', 'Fecha/Hora', 'Medicamento', 'Presentación', 'Componentes', 'Lote', 'Cantidad', 'Cédula Ben.', 'Carga', 'Titular', 'C.I. Titular', 'Farmacéutico']
                titulo = f"Registro de Despachos — {fecha_desde_fmt} al {fecha_hasta_fmt}"
                filename = f"despachos_{fecha_desde_fmt}_{fecha_hasta_fmt}"

        if formato == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename=\"{filename}.csv\"'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.write('\ufeff')  # BOM para Excel
            writer = csv.writer(response)
            writer.writerow(columns)
            for row in rows:
                writer.writerow([str(c) for c in row])
            
            log_descarga(request, f"ESTADISTICAS_CSV_{tipo.upper()}", f"{filename}.csv")
            return response

        elif formato == 'excel':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                wb = Workbook()
                ws = wb.active
                ws.title = "Reporte"

                # Encabezados con estilo
                header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

                for row_idx, row in enumerate(rows, start=2):
                    for col_idx, value in enumerate(row, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=str(value))

                # Ajustar anchos
                for column in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in column)
                    ws.column_dimensions[column[0].column_letter].width = min(max_len + 4, 50)

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                response = HttpResponse(
                    output.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename=\"{filename}.xlsx\"'
                response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                
                log_descarga(request, f"ESTADISTICAS_EXCEL_{tipo.upper()}", f"{filename}.xlsx")
                return response
            except ImportError:
                return Response({'detail': 'openpyxl no está instalado.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        elif formato == 'pdf':
            try:
                from reportlab.lib.pagesizes import letter, landscape
                from reportlab.lib import colors
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib.enums import TA_CENTER

                buffer = io.BytesIO()
                # Use letter and reduce margins to maximize space
                doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=10, rightMargin=10, topMargin=20, bottomMargin=20)
                styles = getSampleStyleSheet()
                styles['Title'].alignment = TA_CENTER
                elements = []

                # Membrete institucional (Buscamos primero en static del backend, con fallback al frontend)
                from reportlab.platypus import Table as RLTable, TableStyle as RLTableStyle
                from reportlab.lib.units import cm
                from reportlab.platypus import Image as RLImage
                import base64
                from api.utils import get_logo_base64
                
                logo_left = None
                logo_right = None
                
                # Intentar cargarlos en Base64 primero para producción
                dem_base64 = get_logo_base64('dem-2.png')
                SIFARMA_base64 = get_logo_base64('logo sifarma - version claro.png')
                
                if dem_base64 and SIFARMA_base64:
                    try:
                        logo_left = RLImage(io.BytesIO(base64.b64decode(dem_base64)), width=3.5*cm, height=1.2*cm)
                        logo_right = RLImage(io.BytesIO(base64.b64decode(SIFARMA_base64)), width=5.8*cm, height=1.4*cm)
                    except Exception:
                        pass
                
                # Fallback al sistema de archivos local si falla Base64
                if not logo_left or not logo_right:
                    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'dem-2.png')
                    if not os.path.exists(logo_path):
                        logo_path = os.path.join(settings.BASE_DIR, '..', 'frontend', 'src', 'assets', 'img', 'dem-2.png')
                        
                    SIFARMA_logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo sifarma - version claro.png')
                    if not os.path.exists(SIFARMA_logo_path):
                        SIFARMA_logo_path = os.path.join(settings.BASE_DIR, '..', 'frontend', 'src', 'assets', 'img', 'logo sifarma - version claro.png')
                    
                    if os.path.exists(logo_path) and os.path.exists(SIFARMA_logo_path):
                        try:
                            logo_left = RLImage(logo_path, width=3.5*cm, height=1.2*cm)
                            logo_right = RLImage(SIFARMA_logo_path, width=5.8*cm, height=1.4*cm)
                        except Exception:
                            pass

                header_data = []
                if logo_left and logo_right:
                    header_data = [[
                        logo_left, 
                        Paragraph(
                            '<div align="center"><b>DIRECCIÓN EJECUTIVA DE LA MAGISTRATURA</b><br/>'
                            '<font size="8">SISTEMA DE INVENTARIO DE FARMACIA - SIFARMA</font><br/>'
                            f'<font size="7" color="grey">Período: {fecha_desde_fmt} al {fecha_hasta_fmt}</font></div>',
                            styles['Normal']
                        ),
                        logo_right
                    ]]
                    header_table = RLTable(header_data, colWidths=['20%', '60%', '20%'])
                    header_table.setStyle(RLTableStyle([
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
                        ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#1e3a5f')),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ]))
                else:
                    header_data = [[Paragraph(
                        '<div align="center"><b>DIRECCIÓN EJECUTIVA DE LA MAGISTRATURA</b><br/>'
                        f'<font size="8">SIFARMA | Período: {fecha_desde_fmt} al {fecha_hasta_fmt}</font></div>',
                        styles['Normal']
                    )]]
                    header_table = RLTable(header_data, colWidths=['100%'])
                    header_table.setStyle(RLTableStyle([
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#1e3a5f')),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ]))
                elements.append(header_table)
                elements.append(Spacer(1, 6))
                elements.append(Paragraph(f'<b>{titulo}</b>', styles['Title']))
                elements.append(Spacer(1, 8))

                if tipo == 'inventario':
                    col_widths = [60, 100, 100, 190, 60, 60, 60, 60]
                elif tipo == 'top_medicamentos':
                    col_widths = [140, 140, 290, 100, 100]
                elif tipo == 'ingresos':
                    col_widths = [70, 100, 100, 160, 60, 80, 70, 60, 70]
                else: # despachos
                    col_widths = [50, 65, 80, 80, 120, 50, 35, 50, 30, 80, 50, 60]

                table_data = [columns]
                for row in rows:
                    row_data = []
                    for c in row:
                        val = str(c) if c is not None else ''
                        row_data.append(Paragraph(f'<font size="6">{val}</font>', styles['Normal']))
                    table_data.append(row_data)

                t = Table(table_data, colWidths=col_widths, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(t)
                doc.build(elements)

                buffer.seek(0)
                response = HttpResponse(buffer.read(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
                response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
                
                log_descarga(request, f"ESTADISTICAS_PDF_{tipo.upper()}", f"{filename}.pdf")
                return response
            except ImportError:
                return Response({'detail': 'reportlab no está instalado.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({'detail': 'Formato no soportado. Use: csv, excel, pdf'}, status=status.HTTP_400_BAD_REQUEST)

class ReporteIngresosDetalleView(APIView):
    """Listado detallado de ingresos (dotaciones) en un periodo."""
    permission_classes = [IsStatsViewer]

    def get(self, request):
        fecha_desde, fecha_hasta = get_date_range(request)
        ordering = request.query_params.get('ordering', 'fecha_ingreso').strip()
        
        # Mapeo de campos permitidos para ordenar
        order_map = {
            'fecha_ingreso': 'l.fecha_ingreso DESC',
            'nombre_generico': 'm.nombre_generico ASC',
            'cantidad': 'l.cantidad_inicial DESC',
            'lote': 'l.numero_lote ASC'
        }
        order_sql = order_map.get(ordering, 'l.fecha_ingreso DESC')

        schemas = _get_schemas(request)
        q_list = []
        for sch in _get_valid_schemas(schemas, 'lotes'):
            q_list.append(f"""
                SELECT 
                    l.id_lote, l.numero_lote, m.nombre_generico,
                    (
                        SELECT string_agg(UPPER(pa.nombre_principio || ' ' || mc.concentracion_valor || ' ' || u.nombre_unidad), ' - ')
                        FROM farmacia.medicamento_componentes mc
                        JOIN farmacia.principios_activos pa ON mc.id_principio = pa.id_principio
                        JOIN farmacia.unidades_medida u ON mc.id_unidad = u.id_unidad
                        WHERE mc.id_med_base = m.id_med_base
                    ) AS componentes,
                    p.nombre_presentacion, l.cantidad_inicial,
                    ((l.fecha_ingreso AT TIME ZONE 'UTC') AT TIME ZONE 'America/Caracas') AS fecha_ingreso,
                    au.username AS usuario,
                    CASE 
                        WHEN al.descripcion ILIKE '%%Carga masiva%%' THEN 'MASIVO'
                        ELSE 'MANUAL'
                    END AS tipo_carga
                FROM {sch}.lotes l
                JOIN farmacia.medicamentos_base m ON l.id_med_base = m.id_med_base
                LEFT JOIN farmacia.presentaciones_medicamento p ON m.id_presentacion = p.id_presentacion
                LEFT JOIN public.auth_user au ON l.usuario_registro = au.id
                LEFT JOIN public.auditoria_logs al ON al.accion IN ('INCLUSION', 'INVENTARIO_INCLUSION') 
                     AND al.descripcion ILIKE '%%' || l.numero_lote || '%%'
                WHERE DATE((l.fecha_ingreso AT TIME ZONE 'UTC') AT TIME ZONE 'America/Caracas') BETWEEN %s AND %s
            """)

        if not q_list:
            sql = "SELECT 1 WHERE 1=0"
            params = []
        else:
            sql = f"SELECT * FROM ( {' UNION ALL '.join(q_list)} ) sub ORDER BY {order_sql.replace('l.fecha_ingreso', 'fecha_ingreso').replace('m.nombre_generico', 'nombre_generico').replace('l.cantidad_inicial', 'cantidad_inicial').replace('l.numero_lote', 'numero_lote')}"
            params = [fecha_desde, fecha_hasta] * len(_get_valid_schemas(schemas, 'lotes'))
        
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        return Response(data)




